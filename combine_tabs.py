from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string

def combine_tabs(path):
    # Load the input workbook
    input_wb = load_workbook(filename=path, read_only=True)

    # Create a new workbook and worksheet
    output_wb = Workbook()
    output_ws = output_wb.active

    # Copy data from each worksheet to the new worksheet
    for i, ws in enumerate(input_wb.worksheets):
        # Add the title row
        title = f"Table {i+1}"
        output_ws.append([title])
        output_ws.cell(row=1, column=1).font = Font(bold=True)

        # Copy the data from the worksheet
        for row in ws.iter_rows():
            new_row = []
            for cell in row:
                new_row.append(cell.value)
            output_ws.append(new_row)

    # Save the output workbook
    output_wb.save(filename=path)

